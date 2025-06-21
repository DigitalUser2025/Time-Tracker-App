from flask import render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
from app import app, db
from models import User, TimeEntry, ManualTimeEntry, DailyLog, PasswordResetRequest, calculate_daily_hours
import logging
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def is_logged_in():
    return 'user_id' in session

def current_user():
    if is_logged_in():
        return User.query.get(session['user_id'])
    return None

def requires_auth(func):
    def wrapper(*args, **kwargs):
        if not is_logged_in():
            return redirect(url_for('login'))
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper

def requires_admin_or_manager(func):
    def wrapper(*args, **kwargs):
        if not is_logged_in():
            return redirect(url_for('login'))
        user = current_user()
        if user.role not in ['admin', 'manager']:
            flash('Access denied. Admin or Manager role required.', 'error')
            return redirect(url_for('dashboard'))
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper

def requires_admin(func):
    def wrapper(*args, **kwargs):
        if not is_logged_in():
            return redirect(url_for('login'))
        user = current_user()
        if user.role != 'admin':
            flash('Access denied. Admin role required.', 'error')
            return redirect(url_for('dashboard'))
        return func(*args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper

@app.route('/')
def index():
    if is_logged_in():
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        phone_number = request.form['phone_number']
        password = request.form['password']
        
        user = User.query.filter_by(phone_number=phone_number).first()
        
        if user and check_password_hash(user.password_hash, password):
            if not user.is_approved:
                flash('Your account is pending approval.', 'warning')
                return render_template('login.html')
            
            session['user_id'] = user.id
            session['user_role'] = user.role
            
            # Set persistent session for employees
            if user.role == 'employee':
                session.permanent = True
            
            flash(f'Welcome back, {user.full_name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid phone number or password.', 'error')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        phone_number = request.form['phone_number']
        password = request.form['password']
        full_name = request.form['full_name']
        security_question = request.form['security_question']
        security_answer = request.form['security_answer']
        
        # Check if user already exists
        existing_user = User.query.filter_by(phone_number=phone_number).first()
        if existing_user:
            flash('Phone number already registered.', 'error')
            return render_template('register.html')
        
        # Check if this is the first user (auto-admin)
        user_count = User.query.count()
        is_first_user = user_count == 0
        
        user = User(
            phone_number=phone_number,
            password_hash=generate_password_hash(password),
            full_name=full_name,
            security_question=security_question,
            security_answer=security_answer.lower().strip(),
            role='admin' if is_first_user else 'employee',
            is_approved=is_first_user
        )
        
        db.session.add(user)
        db.session.commit()
        
        if is_first_user:
            flash('Registration successful! You are now the administrator.', 'success')
            session['user_id'] = user.id
            session['user_role'] = user.role
            return redirect(url_for('dashboard'))
        else:
            flash('Registration successful! Your account is pending approval.', 'success')
            return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@requires_auth
def dashboard():
    user = current_user()
    today = date.today()
    
    # Get current month's total hours
    first_day = today.replace(day=1)
    current_month_logs = DailyLog.query.filter_by(user_id=user.id).filter(
        DailyLog.date >= first_day,
        DailyLog.date <= today
    ).all()
    
    current_month_hours = sum(log.total_hours for log in current_month_logs)
    
    # Get recent daily logs (last 7 days)
    seven_days_ago = today - timedelta(days=7)
    recent_logs = DailyLog.query.filter_by(user_id=user.id).filter(
        DailyLog.date >= seven_days_ago
    ).order_by(DailyLog.date.desc()).all()
    
    # Check if user is currently clocked in
    current_entry = TimeEntry.query.filter_by(
        user_id=user.id,
        date=today,
        clock_out=None
    ).first()
    
    # Get pending approvals count for managers/admins
    pending_users = 0
    pending_time_entries = 0
    pending_password_resets = 0
    
    if user.role in ['admin', 'manager']:
        pending_users = User.query.filter_by(is_approved=False).count()
        pending_time_entries = ManualTimeEntry.query.filter_by(status='pending').count()
        pending_password_resets = PasswordResetRequest.query.filter_by(status='pending').count()
    
    return render_template('dashboard.html',
                         user=user,
                         current_month_hours=current_month_hours,
                         recent_logs=recent_logs,
                         is_clocked_in=current_entry is not None,
                         current_entry=current_entry,
                         pending_users=pending_users,
                         pending_time_entries=pending_time_entries,
                         pending_password_resets=pending_password_resets,
                         today=today)

@app.route('/clock_in', methods=['POST'])
@requires_auth
def clock_in():
    user = current_user()
    today = date.today()
    now = datetime.now()
    
    # Check if already clocked in
    current_entry = TimeEntry.query.filter_by(
        user_id=user.id,
        date=today,
        clock_out=None
    ).first()
    
    if current_entry:
        flash('You are already clocked in!', 'warning')
    else:
        entry = TimeEntry(
            user_id=user.id,
            clock_in=now,
            date=today
        )
        db.session.add(entry)
        db.session.commit()
        flash('Clocked in successfully!', 'success')
    
    return redirect(url_for('dashboard'))

@app.route('/clock_out', methods=['POST'])
@requires_auth
def clock_out():
    user = current_user()
    today = date.today()
    now = datetime.now()
    
    # Find current open entry
    current_entry = TimeEntry.query.filter_by(
        user_id=user.id,
        date=today,
        clock_out=None
    ).first()
    
    if current_entry:
        current_entry.clock_out = now
        
        # Calculate hours worked
        delta = now - current_entry.clock_in
        hours = delta.total_seconds() / 3600
        current_entry.hours_worked = hours
        
        db.session.commit()
        
        # Update daily log
        calculate_daily_hours(user.id, today)
        
        flash(f'Clocked out successfully! Worked {hours:.2f} hours.', 'success')
    else:
        flash('You are not currently clocked in.', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/manual_time_entry', methods=['GET', 'POST'])
@requires_auth
def manual_time_entry():
    if request.method == 'POST':
        user = current_user()
        entry_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        hours = float(request.form['hours'])
        reason = request.form['reason']
        
        # Check if entry already exists for this date
        existing_entry = ManualTimeEntry.query.filter_by(
            user_id=user.id,
            date=entry_date
        ).first()
        
        if existing_entry:
            flash('Manual time entry already exists for this date.', 'error')
        else:
            entry = ManualTimeEntry(
                user_id=user.id,
                date=entry_date,
                hours=hours,
                reason=reason
            )
            db.session.add(entry)
            db.session.commit()
            flash('Manual time entry submitted for approval.', 'success')
            return redirect(url_for('dashboard'))
    
    today = date.today()
    return render_template('manual_time_entry.html', today=today)

@app.route('/reports')
@requires_admin_or_manager
def reports():
    user = current_user()
    
    # Get filter parameters
    employee_id = request.args.get('employee_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Build query
    query = DailyLog.query.join(User)
    
    if employee_id:
        query = query.filter(DailyLog.user_id == employee_id)
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        query = query.filter(DailyLog.date >= start_date)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        query = query.filter(DailyLog.date <= end_date)
    
    logs = query.order_by(DailyLog.date.desc()).all()
    
    # Calculate totals
    total_hours = sum(log.total_hours for log in logs)
    total_regular_hours = sum(log.regular_hours for log in logs)
    total_manual_hours = sum(log.manual_hours for log in logs)
    
    # Get all employees for filter dropdown
    employees = User.query.filter_by(is_approved=True).order_by(User.full_name).all()
    
    return render_template('reports.html',
                         logs=logs,
                         employees=employees,
                         total_hours=total_hours,
                         total_regular_hours=total_regular_hours,
                         total_manual_hours=total_manual_hours,
                         selected_employee=employee_id,
                         start_date=start_date.strftime('%Y-%m-%d') if start_date else '',
                         end_date=end_date.strftime('%Y-%m-%d') if end_date else '')

@app.route('/export_reports')
@requires_admin_or_manager
def export_reports():
    # Get filter parameters
    employee_id = request.args.get('employee_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Build query
    query = DailyLog.query.join(User)
    
    if employee_id:
        query = query.filter(DailyLog.user_id == employee_id)
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        query = query.filter(DailyLog.date >= start_date)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        query = query.filter(DailyLog.date <= end_date)
    
    logs = query.order_by(DailyLog.date.desc()).all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ["Employee", "Date", "Regular Hours", "Manual Hours", "Total Hours"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Data rows
    for row, log in enumerate(logs, 2):
        ws.cell(row=row, column=1, value=log.user.full_name)
        ws.cell(row=row, column=2, value=log.date.strftime('%B %d, %Y'))
        ws.cell(row=row, column=3, value=round(log.regular_hours, 2))
        ws.cell(row=row, column=4, value=round(log.manual_hours, 2))
        ws.cell(row=row, column=5, value=round(log.total_hours, 2))
    
    # Add summary row
    if logs:
        summary_row = len(logs) + 3
        ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=summary_row, column=3, value=round(sum(log.regular_hours for log in logs), 2)).font = Font(bold=True)
        ws.cell(row=summary_row, column=4, value=round(sum(log.manual_hours for log in logs), 2)).font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=round(sum(log.total_hours for log in logs), 2)).font = Font(bold=True)
    
    # Auto-adjust column widths
    for col in range(1, 6):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Create file in memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"time_report_{current_date}.xlsx"
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response

@app.route('/admin_panel')
@requires_admin
def admin_panel():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin_panel.html', users=users)

@app.route('/assign_role', methods=['POST'])
@requires_admin
def assign_role():
    user_id = request.form['user_id']
    new_role = request.form['role']
    
    user = User.query.get(user_id)
    if user:
        user.role = new_role
        db.session.commit()
        flash(f'Role updated for {user.full_name}', 'success')
    
    return redirect(url_for('admin_panel'))

@app.route('/approve_users')
@requires_admin_or_manager
def approve_users():
    pending_users = User.query.filter_by(is_approved=False).all()
    return render_template('approve_users.html', pending_users=pending_users)

@app.route('/approve_user/<int:user_id>')
@requires_admin_or_manager
def approve_user(user_id):
    user = User.query.get(user_id)
    if user:
        user.is_approved = True
        db.session.commit()
        flash(f'User {user.full_name} approved successfully.', 'success')
    return redirect(url_for('approve_users'))

@app.route('/reject_user/<int:user_id>')
@requires_admin_or_manager
def reject_user(user_id):
    user = User.query.get(user_id)
    if user:
        db.session.delete(user)
        db.session.commit()
        flash(f'User {user.full_name} rejected and removed.', 'info')
    return redirect(url_for('approve_users'))

@app.route('/delete_user/<int:user_id>')
@requires_admin_or_manager
def delete_user(user_id):
    current_user_obj = current_user()
    user = User.query.get(user_id)
    
    if user and user.id != current_user_obj.id:
        # Delete related records first to avoid foreign key constraints
        TimeEntry.query.filter_by(user_id=user_id).delete()
        ManualTimeEntry.query.filter_by(user_id=user_id).delete()
        ManualTimeEntry.query.filter_by(approved_by=user_id).update({'approved_by': None})
        DailyLog.query.filter_by(user_id=user_id).delete()
        PasswordResetRequest.query.filter_by(user_id=user_id).delete()
        PasswordResetRequest.query.filter_by(approved_by=user_id).update({'approved_by': None})
        
        # Delete the user
        db.session.delete(user)
        db.session.commit()
        flash(f'User {user.full_name} and all related data have been permanently deleted.', 'info')
    elif user and user.id == current_user_obj.id:
        flash('You cannot delete your own account.', 'error')
    else:
        flash('User not found.', 'error')
    
    return redirect(url_for('admin_panel'))

@app.route('/approve_time_entries')
@requires_admin_or_manager
def approve_time_entries():
    pending_entries = ManualTimeEntry.query.filter_by(status='pending').all()
    return render_template('approve_time_entries.html', pending_entries=pending_entries)

@app.route('/approve_time_entry/<int:entry_id>')
@requires_admin_or_manager
def approve_time_entry(entry_id):
    entry = ManualTimeEntry.query.get(entry_id)
    if entry:
        entry.status = 'approved'
        entry.approved_by = current_user().id
        entry.approved_at = datetime.utcnow()
        db.session.commit()
        
        # Recalculate daily hours
        calculate_daily_hours(entry.user_id, entry.date)
        
        flash(f'Time entry approved for {entry.user.full_name}.', 'success')
    return redirect(url_for('approve_time_entries'))

@app.route('/reject_time_entry/<int:entry_id>')
@requires_admin_or_manager
def reject_time_entry(entry_id):
    entry = ManualTimeEntry.query.get(entry_id)
    if entry:
        entry.status = 'rejected'
        entry.approved_by = current_user().id
        entry.approved_at = datetime.utcnow()
        db.session.commit()
        flash(f'Time entry rejected for {entry.user.full_name}.', 'info')
    return redirect(url_for('approve_time_entries'))

@app.route('/edit_time_entry')
@requires_admin_or_manager
def edit_time_entry():
    # Get filter parameters
    employee_id = request.args.get('employee_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Build query for time entries
    query = TimeEntry.query.join(User)
    
    if employee_id:
        query = query.filter(TimeEntry.user_id == employee_id)
    
    if start_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        query = query.filter(TimeEntry.date >= start_date_obj)
    
    if end_date:
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        query = query.filter(TimeEntry.date <= end_date_obj)
    
    time_entries = query.order_by(TimeEntry.date.desc(), TimeEntry.clock_in.desc()).all()
    
    # Get all employees for filter dropdown
    employees = User.query.filter_by(is_approved=True, role='employee').order_by(User.full_name).all()
    
    return render_template('edit_time_entry.html',
                         time_entries=time_entries,
                         employees=employees,
                         selected_employee=employee_id,
                         start_date=start_date or '',
                         end_date=end_date or '')

@app.route('/update_time_entry/<int:entry_id>', methods=['POST'])
@requires_admin_or_manager
def update_time_entry(entry_id):
    entry = TimeEntry.query.get(entry_id)
    if not entry:
        flash('Time entry not found.', 'error')
        return redirect(url_for('edit_time_entry'))
    
    try:
        # Get form data
        clock_in_str = request.form['clock_in']
        clock_out_str = request.form['clock_out']
        
        # Parse datetime strings
        clock_in = datetime.strptime(clock_in_str, '%Y-%m-%dT%H:%M')
        clock_out = datetime.strptime(clock_out_str, '%Y-%m-%dT%H:%M') if clock_out_str else None
        
        # Validate times
        if clock_out and clock_out <= clock_in:
            flash('Clock out time must be after clock in time.', 'error')
            return redirect(url_for('edit_time_entry'))
        
        # Update entry
        entry.clock_in = clock_in
        entry.clock_out = clock_out
        entry.date = clock_in.date()
        
        # Calculate hours worked
        if clock_out:
            delta = clock_out - clock_in
            entry.hours_worked = delta.total_seconds() / 3600
        else:
            entry.hours_worked = 0.0
        
        db.session.commit()
        
        # Recalculate daily hours
        calculate_daily_hours(entry.user_id, entry.date)
        
        flash(f'Time entry updated for {entry.user.full_name}.', 'success')
        
    except ValueError as e:
        flash('Invalid date/time format. Please use the correct format.', 'error')
    except Exception as e:
        flash('An error occurred while updating the time entry.', 'error')
        db.session.rollback()
    
    return redirect(url_for('edit_time_entry'))

@app.route('/delete_time_entry/<int:entry_id>')
@requires_admin_or_manager
def delete_time_entry(entry_id):
    entry = TimeEntry.query.get(entry_id)
    if not entry:
        flash('Time entry not found.', 'error')
        return redirect(url_for('edit_time_entry'))
    
    user_id = entry.user_id
    entry_date = entry.date
    user_name = entry.user.full_name
    
    db.session.delete(entry)
    db.session.commit()
    
    # Recalculate daily hours
    calculate_daily_hours(user_id, entry_date)
    
    flash(f'Time entry deleted for {user_name}.', 'success')
    return redirect(url_for('edit_time_entry'))

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        phone_number = request.form['phone_number']
        security_answer = request.form['security_answer'].lower().strip()
        
        user = User.query.filter_by(phone_number=phone_number).first()
        
        if user and user.security_answer == security_answer:
            # Correct security answer, allow direct password reset
            session['reset_user_id'] = user.id
            return redirect(url_for('reset_password'))
        elif user:
            # Incorrect security answer, create reset request
            existing_request = PasswordResetRequest.query.filter_by(
                user_id=user.id,
                status='pending'
            ).first()
            
            if not existing_request:
                reset_request = PasswordResetRequest(user_id=user.id)
                db.session.add(reset_request)
                db.session.commit()
            
            flash('Security answer incorrect. A password reset request has been sent to administrators for approval.', 'info')
            return redirect(url_for('login'))
        else:
            flash('Phone number not found.', 'error')
    
    return render_template('forgot_password.html')

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if 'reset_user_id' not in session:
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        new_password = request.form['password']
        user = User.query.get(session['reset_user_id'])
        
        if user:
            user.password_hash = generate_password_hash(new_password)
            db.session.commit()
            session.pop('reset_user_id', None)
            flash('Password reset successfully. Please log in with your new password.', 'success')
            return redirect(url_for('login'))
    
    return render_template('reset_password.html')

@app.route('/approve_password_reset/<int:request_id>')
@requires_admin_or_manager
def approve_password_reset(request_id):
    reset_request = PasswordResetRequest.query.get(request_id)
    if reset_request:
        reset_request.status = 'approved'
        reset_request.approved_by = current_user().id
        reset_request.approved_at = datetime.utcnow()
        db.session.commit()
        
        # Here you would typically send a notification or email
        flash(f'Password reset approved for {reset_request.user.full_name}.', 'success')
    
    return redirect(url_for('dashboard'))

@app.route('/filter_dashboard', methods=['POST'])
@requires_auth
def filter_dashboard():
    user = current_user()
    start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date()
    
    logs = DailyLog.query.filter_by(user_id=user.id).filter(
        DailyLog.date >= start_date,
        DailyLog.date <= end_date
    ).order_by(DailyLog.date.desc()).all()
    
    total_hours = sum(log.total_hours for log in logs)
    
    return render_template('dashboard.html',
                         user=user,
                         filtered_logs=logs,
                         filtered_total_hours=total_hours,
                         filter_start_date=start_date,
                         filter_end_date=end_date)
